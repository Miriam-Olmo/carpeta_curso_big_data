

# Importación del módulo nativo para procesar archivos separados por comas (CSV)
import csv
# Importación de la librería estándar para Serialización/Deserialización de objetos JSON
import json
# Importación de la API de ElementTree para analizar y construir estructuras jerárquicas XML
import xml.etree.ElementTree as et
# Importación de la función específica de openpyxl para abrir libros de trabajo Excel en memoria
from openpyxl import load_workbook


def cargar_csv(ruta):
    """Abre un archivo CSV y mapea automáticamente sus filas a diccionarios en memoria."""
    # Abre el descriptor de archivo en modo lectura de texto r, forzando codificación UTF-8 para tildes
    fichero = open(ruta, "r", encoding='UTF-8')
    # DictReader mapea las celdas de cada fila usando como llaves los textos de la primera línea de cabeceras
    lector = csv.DictReader(fichero) 
    # Transforma el iterador perezoso en una estructura física de lista con todos los diccionarios cargados
    artistas = list(lector)
    # Cierra el flujo del archivo para liberar los recursos del sistema operativo de inmediato
    fichero.close()
    # Retorna la lista homogénea de registros/filas procesada
    return artistas


def cargar_excel(ruta):
    """Lee la pestaña activa de un libro Excel y la transforma en una lista de diccionarios."""
    # Instancia el objeto Workbook de openpyxl mapeando el archivo binario de la ruta dada
    excel = load_workbook(ruta)
    # Captura la primera hoja de cálculo disponible o la que quedó seleccionada al guardar el archivo
    hoja = excel.active
    
    # Crea un generador iterador que extrae los datos puros sin formatos estéticos (values_only=True)
    filas = hoja.iter_rows(values_only=True)
    # Extrae el primer elemento del generador (fila 1), el cual contiene los nombres de las columnas
    cabeceras = next(filas)
    
    # Inicializa una estructura de lista vacía que contendrá los diccionarios resultantes
    lista_resultante = []
    # Itera fila por fila omitiendo la fila 1 (cabeceras) gracias al parámetro min_row=2
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        # zip combina en tuplas la cabecera con su respectivo valor de celda; dict() lo vuelve un diccionario
        producto = dict(zip(cabeceras, fila))
        # Añade el nuevo registro estructurado a la lista de almacenamiento
        lista_resultante.append(producto)
       
    # Retorna la colección completa de registros extraídos del Excel
    return lista_resultante


def cargar_json(ruta):
    """Parsea archivos JSON estructurados transformándolos en diccionarios y listas de Python."""
    # Abre el archivo origen en modo de lectura con soporte de caracteres internacionales
    fichero = open(ruta, "r", encoding="UTF-8")
    # json.load decodifica directamente el flujo de texto plano del archivo a tipos nativos de Python
    datos = json.load(fichero)
    # Cierra el archivo liberando el puntero del disco duro
    fichero.close()
    # Devuelve los datos deserializados (generalmente una lista de diccionarios)
    return datos


def cargar_xml(ruta):
    """Parsea un archivo jerárquico XML extrayendo los nodos hijos en una lista plana."""
    # et.parse analiza de forma secuencial la estructura de etiquetas del archivo en disco
    arbol = et.parse(ruta)
    # getroot apunta directamente al nodo contenedor del nivel superior (ej: <patrocinadores>)
    nodo_raiz = arbol.getroot()

    # Inicializa la lista receptora donde unificaremos el formato a diccionario común
    patrocinadores = []
    # Itera de forma manual por cada etiqueta hija exacta denominada '<patrocinador>'
    for patrocinador in nodo_raiz.findall('patrocinador'):
        # Mapeo manual: Construye un diccionario buscando las subetiquetas internas (.find) y extrayendo su texto plano (.text)
        datos_patrocinador = {
            'nombre_empresa': patrocinador.find('nombre_empresa').text,
            'contacto': patrocinador.find('contacto').text,
            # Estructura de control condicional (inline if) para evitar errores si la etiqueta opcional viniera ausente (None)
            'email': patrocinador.find('email').text if patrocinador.find('email') is not None else None,
            'categoria': patrocinador.find('categoria').text if patrocinador.find('categoria') is not None else None,
            'monto_patrocinio': patrocinador.find('monto_patrocinio').text if patrocinador.find('monto_patrocinio') is not None else None
        }
        # Agrega el diccionario plano recién construido a la lista final
        patrocinadores.append(datos_patrocinador)
        
    # Retorna el conjunto homogéneo de patrocinadores listo para el pipeline de limpieza
    return patrocinadores